"""Импорт внутренних реестров Разноторга «для ИИ».

Поддерживает отдельные xlsx, директории и zip-архивы. По умолчанию внутри zip
берутся только файлы с названием вида «Реестр ... для ИИ.xlsx», чтобы не
перебирать тяжёлые дневные заказы из почтовых архивов.
"""
from __future__ import annotations

import argparse
import re
import sys
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db import SessionLocal, init_db  # noqa: E402
from app.models import RetailHistoryItem  # noqa: E402


HEADER_SCAN_ROWS = 50
PERIOD_RE = re.compile(r"(\d{2}\.\d{2}\.\d{4})\s*[-–]\s*(\d{2}\.\d{2}\.\d{4})")
DATE_RE = re.compile(r"\d{2}\.\d{2}[.\s]\d{4}")


@dataclass(frozen=True)
class XlsxSource:
    display_name: str
    opener: object


def _norm(v) -> str:
    if v is None:
        return ""
    return " ".join(str(v).strip().lower().split())


def _clean(v) -> str | None:
    if v is None:
        return None
    s = " ".join(str(v).strip().split())
    if not s or s in {"0", "-", "#N/A"}:
        return None
    return s


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\u00a0", " ").replace(" ", "").replace(",", ".").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None


def _looks_like_retail_registry(name: str) -> bool:
    n = name.casefold()
    return n.endswith(".xlsx") and "реестр" in n and "для ии" in n and not Path(name).name.startswith("~$")


def iter_sources(paths: list[Path], *, all_xlsx: bool = False) -> Iterable[XlsxSource]:
    for path in paths:
        if path.is_dir():
            for child in sorted(path.rglob("*.xlsx")):
                if all_xlsx or _looks_like_retail_registry(child.name):
                    yield XlsxSource(child.name, child)
        elif path.suffix.lower() == ".zip":
            with zipfile.ZipFile(path) as zf:
                for info in zf.infolist():
                    if info.is_dir():
                        continue
                    name = Path(info.filename).name
                    if not name.lower().endswith(".xlsx") or name.startswith("~$"):
                        continue
                    if not all_xlsx and not _looks_like_retail_registry(name):
                        continue
                    data = zf.read(info)
                    yield XlsxSource(name, BytesIO(data))
        elif path.suffix.lower() == ".xlsx":
            if all_xlsx or _looks_like_retail_registry(path.name):
                yield XlsxSource(path.name, path)


def _load_workbook(src: XlsxSource):
    opener = src.opener
    if isinstance(opener, Path):
        return openpyxl.load_workbook(opener, read_only=True, data_only=True)
    if isinstance(opener, BytesIO):
        opener.seek(0)
        return openpyxl.load_workbook(opener, read_only=True, data_only=True)
    if hasattr(opener, "read"):
        return openpyxl.load_workbook(opener, read_only=True, data_only=True)
    raise TypeError(f"unsupported source: {src.display_name}")


def _find_header(buffer: list[tuple[int, tuple]]) -> tuple[int, dict] | None:
    for row_no, row in buffer:
        cmap = _build_column_map(row)
        required = {"group", "subgroup", "vid", "product_name", "sales_current"}
        if required.issubset(cmap):
            return row_no, cmap
    return None


def _build_column_map(header: tuple) -> dict:
    cmap: dict[str, int] = {}
    profitability_cols: list[int] = []
    sales_cols: list[int] = []
    model_cols: list[int] = []
    for idx, cell in enumerate(header):
        name = _norm(cell)
        if not name:
            continue
        if name in {"группа", "гр"}:
            cmap["group"] = idx
        elif name in {"подгруппа", "пгр"}:
            cmap["subgroup"] = idx
        elif name == "вид":
            cmap["vid"] = idx
        elif name == "ценовая":
            cmap["price_band"] = idx
        elif name == "товар":
            cmap["product_name"] = idx
        elif name == "цена":
            cmap["price"] = idx
        elif "наценк" in name:
            cmap["markup"] = idx
        elif "рентаб" in name:
            profitability_cols.append(idx)
        elif "продажи за период" in name:
            sales_cols.append(idx)
        elif name.startswith("остаток") or name.startswith("остаток на"):
            cmap["stock_current"] = idx
        elif "количество моделей" in name or name.startswith("модели на"):
            model_cols.append(idx)

    if profitability_cols:
        cmap["profitability_current"] = profitability_cols[0]
    if len(profitability_cols) > 1:
        cmap["profitability_prev"] = profitability_cols[1]
    if sales_cols:
        cmap["sales_current"] = sales_cols[0]
    if len(sales_cols) > 1:
        cmap["sales_prev"] = sales_cols[1]
    if model_cols:
        cmap["model_count_current"] = model_cols[0]
    if len(model_cols) > 1:
        cmap["model_count_prev"] = model_cols[1]
    return cmap


def _cell(row: tuple, idx: int | None):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _header_period(header: tuple, idx: int | None) -> str | None:
    text = str(_cell(header, idx) or "")
    m = PERIOD_RE.search(text)
    return f"{m.group(1)} - {m.group(2)}" if m else None


def _header_date(header: tuple, idx: int | None) -> str | None:
    text = str(_cell(header, idx) or "")
    m = DATE_RE.search(text)
    return m.group(0).replace(" ", ".") if m else None


def _is_total(*values: str | None) -> bool:
    return any("итог" in (v or "").casefold() for v in values)


def iter_records(src: XlsxSource) -> Iterable[dict]:
    wb = _load_workbook(src)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows_iter = ws.iter_rows(values_only=True)
        buffer: list[tuple[int, tuple]] = []
        for row_no, row in zip(range(1, HEADER_SCAN_ROWS + 1), rows_iter):
            buffer.append((row_no, tuple(row)))
        found = _find_header(buffer)
        if found is None:
            continue
        header_no, cmap = found
        header = dict(buffer)[header_no]
        period_current = _header_period(header, cmap.get("sales_current"))
        period_prev = _header_period(header, cmap.get("sales_prev"))
        stock_date = _header_date(header, cmap.get("stock_current"))

        context = {"group": None, "subgroup": None, "vid": None, "price_band": None}
        data_rows = [row for row_no, row in buffer if row_no > header_no]
        data_rows.extend(tuple(row) for row in rows_iter)
        for row in data_rows:
            raw_group = _clean(_cell(row, cmap.get("group")))
            raw_subgroup = _clean(_cell(row, cmap.get("subgroup")))
            raw_vid = _clean(_cell(row, cmap.get("vid")))
            raw_price_band = _clean(_cell(row, cmap.get("price_band")))
            product_name = _clean(_cell(row, cmap.get("product_name")))
            for key, value in (
                ("group", raw_group),
                ("subgroup", raw_subgroup),
                ("vid", raw_vid),
                ("price_band", raw_price_band),
            ):
                if value and not _is_total(value):
                    context[key] = value

            group = context["group"]
            subgroup = context["subgroup"]
            vid = context["vid"]
            price_band = context["price_band"]
            if not group or not subgroup or not vid or not product_name:
                continue
            if _is_total(group, subgroup, vid, price_band, product_name):
                continue
            if _norm(product_name) == "товар":
                continue

            yield {
                "group": group,
                "subgroup": subgroup,
                "vid": vid,
                "price_band": price_band,
                "product_name": product_name,
                "price": _to_float(_cell(row, cmap.get("price"))),
                "markup": _to_float(_cell(row, cmap.get("markup"))),
                "profitability_current": _to_float(_cell(row, cmap.get("profitability_current"))),
                "profitability_prev": _to_float(_cell(row, cmap.get("profitability_prev"))),
                "sales_current": _to_float(_cell(row, cmap.get("sales_current"))) or 0.0,
                "sales_prev": _to_float(_cell(row, cmap.get("sales_prev"))) or 0.0,
                "stock_current": _to_float(_cell(row, cmap.get("stock_current"))) or 0.0,
                "model_count_current": _to_float(_cell(row, cmap.get("model_count_current"))),
                "model_count_prev": _to_float(_cell(row, cmap.get("model_count_prev"))),
                "period_current": period_current,
                "period_prev": period_prev,
                "stock_date": stock_date,
                "source_file": src.display_name,
                "source_sheet": sheet,
            }


def import_sources(paths: list[Path], *, all_xlsx: bool = False, dry_run: bool = False) -> dict:
    init_db()
    stats = {"files": 0, "rows_read": 0, "inserted": 0, "skipped_files": 0}
    for src in iter_sources(paths, all_xlsx=all_xlsx):
        records = list(iter_records(src))
        if not records:
            stats["skipped_files"] += 1
            continue
        stats["files"] += 1
        stats["rows_read"] += len(records)
        if dry_run:
            continue
        with SessionLocal() as db:
            db.query(RetailHistoryItem).filter_by(source_file=src.display_name).delete()
            db.add_all(RetailHistoryItem(**r) for r in records)
            db.commit()
        stats["inserted"] += len(records)
    return stats


def main():
    p = argparse.ArgumentParser()
    p.add_argument("sources", nargs="+", help="xlsx, директория или zip")
    p.add_argument("--all-xlsx", action="store_true", help="пытаться читать все xlsx, а не только 'Реестр ... для ИИ'")
    p.add_argument("--dry-run", action="store_true", help="распарсить без записи в БД")
    args = p.parse_args()
    paths = [Path(s).expanduser().resolve() for s in args.sources]
    missing = [str(p) for p in paths if not p.exists()]
    if missing:
        sys.exit("не найдено: " + ", ".join(missing))
    stats = import_sources(paths, all_xlsx=args.all_xlsx, dry_run=args.dry_run)
    print(
        f"импорт истории: файлов={stats['files']}, строк={stats['rows_read']}, "
        f"insert={stats['inserted']}, skipped={stats['skipped_files']}"
    )


if __name__ == "__main__":
    main()
