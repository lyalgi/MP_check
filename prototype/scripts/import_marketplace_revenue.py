"""Импорт годовой выручки категорий WB и OZON.

Источник: revenue/wb_revenue_year.xlsx, revenue/ozon_revenue_year.xlsx.
Формат каждого: 2 колонки — «Категория маркетплейса», «Выручка год».
Один путь может встретиться несколько раз (разные периоды/срезы) — суммируем.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db import SessionLocal, init_db  # noqa: E402
from app.models import MarketplaceCategoryRevenue  # noqa: E402


def _iter_rows(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    first = next(rows, None)  # пропустить header
    for r in rows:
        if not r or len(r) < 2:
            continue
        path, rev = r[0], r[1]
        if not path:
            continue
        try:
            rev_f = float(rev) if rev is not None else 0.0
        except (TypeError, ValueError):
            continue
        yield str(path).strip(), rev_f


def import_marketplace(marketplace: str, xlsx_path: Path) -> dict:
    if not xlsx_path.exists():
        return {"skipped": True, "reason": f"not found: {xlsx_path}"}
    agg: dict[str, float] = {}
    rows_read = 0
    for path, rev in _iter_rows(xlsx_path):
        rows_read += 1
        agg[path] = agg.get(path, 0.0) + rev

    with SessionLocal() as db:
        inserted = updated = 0
        for path, total_rev in agg.items():
            existing = (
                db.query(MarketplaceCategoryRevenue)
                .filter_by(marketplace=marketplace, path=path)
                .one_or_none()
            )
            if existing is None:
                db.add(MarketplaceCategoryRevenue(
                    marketplace=marketplace, path=path, revenue_year=total_rev,
                ))
                inserted += 1
            else:
                existing.revenue_year = total_rev
                updated += 1
        db.commit()
    return {
        "marketplace": marketplace,
        "rows_read": rows_read,
        "unique_paths": len(agg),
        "inserted": inserted,
        "updated": updated,
    }


def main():
    init_db()
    root = Path(__file__).resolve().parent.parent / "revenue"
    for mp, name in [("wb", "wb_revenue_year.xlsx"), ("ozon", "ozon_revenue_year.xlsx")]:
        stats = import_marketplace(mp, root / name)
        print(stats)


if __name__ == "__main__":
    main()
