"""Импорт классификаторов Разноторга в SQLite.

Каждый xlsx устроен по-разному (12/15/23/33 столбца, разные позиции header).
Алгоритм:
  1. для каждого листа ищем строку, где совпадают слова "Группа", "Подгруппа", "Вид"
  2. из этого header достаём индексы нужных колонок и всех WB_* / OZON_*
  3. читаем данные ниже header'а и сливаем по ключу (Группа, Подгруппа, Вид)
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Iterable

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db import SessionLocal, init_db  # noqa: E402
from app.models import TaxonomyItem  # noqa: E402
from app.settings import settings  # noqa: E402


NUM_FIELDS = {
    "sold_qty": ["продано"],
    "stock_qty": ["остаток"],
    "sold_rub": ["сумма продаж"],
    "stock_rub": ["сумма остатка"],
    "cost_sold": ["себ прод", "себестоимость продаж", "себ. прод"],
    "cost_stock": ["себ ост", "себестоимость остатка", "себ. ост"],
}


def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        try:
            return float(str(v).replace(" ", "").replace(",", "."))
        except (ValueError, TypeError):
            return 0.0


def _find_header_row(rows: list[list]) -> int | None:
    for i, row in enumerate(rows):
        normed = [_norm(c) for c in row]
        if "группа" in normed and "подгруппа" in normed and "вид" in normed:
            return i
    return None


def _build_column_map(header: list) -> dict:
    """Маппит логическое имя поля на индекс колонки в header'е."""
    cmap: dict[str, int | list[int]] = {}
    wb_cols: list[int] = []
    ozon_cols: list[int] = []

    for idx, cell in enumerate(header):
        name = _norm(cell)
        if not name:
            continue
        if name == "группа":
            cmap["group"] = idx
        elif name == "подгруппа":
            cmap["subgroup"] = idx
        elif name == "вид":
            cmap["vid"] = idx
        elif "товаровед" in name:
            cmap.setdefault("tovaroved", idx)
        else:
            for field, keys in NUM_FIELDS.items():
                if any(name == k or k in name for k in keys):
                    cmap.setdefault(field, idx)
            if "классификатор wb" in name or name.startswith("wb_") or name == "wb":
                wb_cols.append(idx)
            elif "классификатор ozon" in name or name.startswith("ozon_") or name == "ozon":
                ozon_cols.append(idx)
            elif name in ("категория главная", "категория полностью"):
                # Файл OZON «Отделка» — это OZON-категории, тоже учитываем
                ozon_cols.append(idx)
    cmap["_wb_cols"] = wb_cols
    cmap["_ozon_cols"] = ozon_cols
    return cmap


def _clean_path(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s in {"0", "-", "#N/A"}:
        return None
    return s


def _iter_records(xlsx_path: Path) -> Iterable[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        header_idx = _find_header_row(rows[:8])
        if header_idx is None:
            continue
        header = rows[header_idx]
        cmap = _build_column_map(header)
        if "group" not in cmap or "subgroup" not in cmap or "vid" not in cmap:
            continue
        for raw in rows[header_idx + 1:]:
            group = _clean_path(raw[cmap["group"]]) if cmap["group"] < len(raw) else None
            subgroup = _clean_path(raw[cmap["subgroup"]]) if cmap["subgroup"] < len(raw) else None
            vid = _clean_path(raw[cmap["vid"]]) if cmap["vid"] < len(raw) else None
            if not group or not subgroup or not vid:
                continue
            # пропускаем header-as-data строку
            if group.lower() == "группа":
                continue

            def col(name):
                idx = cmap.get(name)
                if idx is None or idx >= len(raw):
                    return None
                return raw[idx]

            wb_paths = [
                p for i in cmap["_wb_cols"]
                if i < len(raw) and (p := _clean_path(raw[i])) is not None
            ]
            ozon_paths = [
                p for i in cmap["_ozon_cols"]
                if i < len(raw) and (p := _clean_path(raw[i])) is not None
            ]
            yield {
                "group": group,
                "subgroup": subgroup,
                "vid": vid,
                "tovaroved": _clean_path(col("tovaroved")),
                "sold_qty": _to_float(col("sold_qty")),
                "stock_qty": _to_float(col("stock_qty")),
                "sold_rub": _to_float(col("sold_rub")),
                "stock_rub": _to_float(col("stock_rub")),
                "cost_sold": _to_float(col("cost_sold")),
                "cost_stock": _to_float(col("cost_stock")),
                "wb_paths": wb_paths,
                "ozon_paths": ozon_paths,
                "source_file": xlsx_path.name,
                "_sheet": sheet,
            }


def import_dir(directory: Path) -> dict:
    init_db()
    files = sorted(directory.glob("*.xlsx"))
    stats = {"files": 0, "rows_read": 0, "unique_items": 0, "inserted": 0, "updated": 0}

    # Сортируем по имени файла — более новый снапшот выигрывает.
    # В именах xlsx часто закодирована дата (например, «..._02.03.26.xlsx»),
    # лексикографическая сортировка отдаёт более позднюю версию последней.
    files = sorted(files, key=lambda p: p.name)

    aggregated: dict[tuple[str, str, str], dict] = {}
    for xlsx_path in files:
        stats["files"] += 1
        for rec in _iter_records(xlsx_path):
            stats["rows_read"] += 1
            key = (rec["group"], rec["subgroup"], rec["vid"])
            prev = aggregated.get(key)
            if prev is None:
                aggregated[key] = {
                    **rec,
                    "wb_paths": set(rec["wb_paths"]),
                    "ozon_paths": set(rec["ozon_paths"]),
                }
            else:
                # WB/OZON-пути ОБЪЕДИНЯЕМ (разные xlsx могут давать разные срезы маппинга).
                prev["wb_paths"] |= set(rec["wb_paths"])
                prev["ozon_paths"] |= set(rec["ozon_paths"])
                if not prev.get("tovaroved") and rec["tovaroved"]:
                    prev["tovaroved"] = rec["tovaroved"]
                # Численные метрики (продажи/остатки/себестоимость) — БЕРЁМ ПОСЛЕДНИЙ.
                # max() искажал картину: показатели с разных дат смешивались.
                # Заведомо нулевые значения не перетирают непустые.
                for f in ("sold_qty", "stock_qty", "sold_rub", "stock_rub", "cost_sold", "cost_stock"):
                    if rec[f] != 0:
                        prev[f] = rec[f]
                prev["source_file"] = rec["source_file"]  # последний выигрывает

    stats["unique_items"] = len(aggregated)

    with SessionLocal() as db:
        for key, rec in aggregated.items():
            existing = db.query(TaxonomyItem).filter_by(
                group=key[0], subgroup=key[1], vid=key[2]
            ).one_or_none()
            wb_paths = sorted(rec["wb_paths"])
            ozon_paths = sorted(rec["ozon_paths"])
            if existing is None:
                db.add(TaxonomyItem(
                    group=rec["group"],
                    subgroup=rec["subgroup"],
                    vid=rec["vid"],
                    tovaroved=rec["tovaroved"],
                    sold_qty=rec["sold_qty"],
                    stock_qty=rec["stock_qty"],
                    sold_rub=rec["sold_rub"],
                    stock_rub=rec["stock_rub"],
                    cost_sold=rec["cost_sold"],
                    cost_stock=rec["cost_stock"],
                    wb_paths=wb_paths,
                    ozon_paths=ozon_paths,
                    source_file=rec["source_file"],
                ))
                stats["inserted"] += 1
            else:
                existing.wb_paths = sorted(set(existing.wb_paths or []) | set(wb_paths))
                existing.ozon_paths = sorted(set(existing.ozon_paths or []) | set(ozon_paths))
                if not existing.tovaroved and rec["tovaroved"]:
                    existing.tovaroved = rec["tovaroved"]
                for f in ("sold_qty", "stock_qty", "sold_rub", "stock_rub", "cost_sold", "cost_stock"):
                    # Повторный импорт должен отражать последний срез, а не
                    # максимум за всё время: иначе остатки/продажи из разных
                    # дат смешиваются и завышают forecast.
                    if rec[f] != 0:
                        setattr(existing, f, rec[f])
                existing.source_file = rec["source_file"]
                stats["updated"] += 1
        db.commit()

    return stats


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--dir", default=settings.classifiers_dir, help="директория с xlsx")
    args = p.parse_args()

    directory = Path(args.dir).resolve()
    if not directory.exists():
        sys.exit(f"директория не найдена: {directory}")
    stats = import_dir(directory)
    print(
        f"импорт завершён. файлов: {stats['files']}, "
        f"строк прочитано: {stats['rows_read']}, "
        f"уникальных видов: {stats['unique_items']}, "
        f"insert: {stats['inserted']}, update: {stats['updated']}"
    )


if __name__ == "__main__":
    main()
